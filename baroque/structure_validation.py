import csv
import os
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# structure_validation.py runs "validate_structure", which calls for "validate_directory" and "validate_file"
# "validate_directory" uses "parse_baroqueproject", "parse_metadata_export", "compare_ids", and "check_empty_directory"
# "validate_file" uses "check_item_files" and "check_intellectual_groups"
# "check_item_files" uses "check_empty_file" and "check_intellectual_groups" uses "create_intellectual_groups"


def parse_baroqueproject(baroqueproject, level):
    ids = []
    paths = []

    for a in getattr(baroqueproject, level):
        ids.append(a["id"])
        paths.append(a["path"])

    return ids, paths


def parse_metadata_export(metadata_export, level):
    """
    This function parses collection IDs or item IDs from the "DigFile Calc" column in the metadata export file.
    The metadata export file can be either csv or xlsx.
    """
    if not os.path.exists(metadata_export):
        print("ERROR: metadata export does not exist")
        sys.exit()

    collection_ids = []
    items_ids = []

    # File type: csv
    if metadata_export.lower().endswith(".csv"):
        with open(metadata_export, "r", newline="") as f:
            reader = csv.DictReader(f)

            for row in reader:
                item_id = row.get("DigFile Calc").strip()
                items_ids.append(item_id)

    # File type: xlsx
    elif metadata_export.lower().endswith(".xlsx"):
        workbook = load_workbook(metadata_export)
        sheet = workbook.active

        # Get the column letter for the "DigFile Calc" column
        target_col_letter = ""
        for col in sheet["1"]:
            if col.value == "DigFile Calc":
                target_col_letter = get_column_letter(col.column)
        # Set the range for rows (containing Item IDs) in the "DigFile Calc" column
        target_range = target_col_letter + "2:" + target_col_letter + str(sheet.max_row)

        for row in sheet[target_range]:
            for cell in row:
                item_id = cell.value.strip()
                items_ids.append(item_id)

    # File type: neither csv nor xlsx
    else:
        print("ERROR: metadata export is unexpected file type")
        sys.exit()

    # NOTE: Collection IDs are parsed from item IDs
    if level == "collections":
        for i in items_ids:
            collection_id = i.split("-")[0]
            if collection_id not in collection_ids:
                collection_ids.append(collection_id)
        return collection_ids

    elif level == "items":
        return items_ids


def compare_ids(set_a_ids, set_b_ids, level):
    diff_ids = list(set(set_a_ids) - set(set_b_ids)) + list(set(set_b_ids) - set(set_a_ids))
    if len(diff_ids) != 0:
        print("ERROR: following " + level[:-1] + " ids do not match:", diff_ids)


def check_empty_directory(directory_path):
    if len(os.listdir(directory_path)) == 0:
        print("ERROR: " + os.path.basename(directory_path) + " is empty directory")


def check_empty_file(file_path):
    for dirpath, dirnames, filenames in os.walk(file_path):
        for filename in filenames:
            if ".txt" not in filename:
                file_path = os.path.join(dirpath, filename)
                file_size = os.path.getsize(file_path)
                if file_size == 0:
                    print("ERROR: " + os.path.basename(file_path) + " is empty file")


def validate_directory(baroqueproject, metadata_export, level):
    """
    To validate a directory name, this function parses
    (1) IDs from a baroqueproject object (which is based on what the vendor returned) and
    (2) a metadata_export file (which is based on what BHL sent)
    create two ID lists, cross-compare the two ID lists, and print any ID that does not exists NOT in BOTH lists.
    It also checks for any empty directory.
    """
    process_ids, process_paths = parse_baroqueproject(baroqueproject, level)
    export_ids = parse_metadata_export(metadata_export, level)

    compare_ids(process_ids, export_ids, level)

    for path in process_paths:
        check_empty_directory(path)


def check_item_files(baroqueproject):
    """
    Check that the minimum number of required files are present for each item (e.g., 1234-SR-1).
    Check that each item does not have more than the maximum number of permissible files.
    Report out any "other" files and check whether any files are empty.
    """
    # Minimum number of files required.
    min_file_nums = {"wav": 2, "mp3": 1, "md5": 3, "jpg": 3, "xml": 1, "txt": 0}
    # Maximum number of files permissible.
    max_file_nums = {"xml": 1, "txt": 1}

    for item in baroqueproject.items:

        min_output = {}
        max_output = {}
        other_output = []
        name_output = []

        for file_type, files in item["files"].items():
            # For each item, output the file type and the number of files that are below the minimum requirements.
            if file_type != "other":
                if len(files) < min_file_nums.get(file_type):
                    min_output[file_type] = len(files)
            # For each item, output the file type and the number of files that are above the maximum requirements.
            if file_type in max_file_nums.keys():
                if len(files) > max_file_nums.get(file_type):
                    max_output[file_type] = len(files)
            # For each item, output the file type and the number of files that are identified as "other" files.
            if file_type == "other":
                if len(files) > 0:
                    other_output = files
            # For each item, out the file names that do no start with the item id.
            for file in files:
                if not file.startswith(item["id"]):
                    name_output.append(file)

        if len(min_output.keys()) > 0:
            print("ERROR: " + item["id"] + " does not have minimum required files:", min_output)
        if len(max_output.keys()) > 0:
            print("ERROR: " + item["id"] + " has more than maximum required files:", max_output)
        if len(other_output) > 0:
            for file in other_output:
                if file not in ["Thumbs.db", ".DS_Store", "Desktop DB", "Desktop DF"]:
                    print("ERROR: " + item["id"] + " has other files:", other_output)
        if len(name_output) > 0:
            for file in name_output:
                if file not in ["Thumbs.db", ".DS_Store", "Desktop DB", "Desktop DF"]:
                    print("ERROR: " + item["id"] + " has file names that do not start with the item id:", name_output)

        check_empty_file(item["path"])


def create_intellectual_groups(baroqueproject):
    """
    Create intellectual groups of md5, mp3, and wav files. Return a dictionary with the following format:
        {
            item id (e.g., 85429-SR-1) :
            {
                part id (e.g., 85429-SR-1-1) : [part files (e.g., 85429-SR-1-1-am.wav)]
            }
        }
    """
    part_file_types = ["md5", "mp3", "wav"]
    items = {}

    for item in baroqueproject.items:
        for file_type, files in item["files"].items():
            if file_type in part_file_types:
                for file in files:

                    # Isolate file names without extensions (e.g., "-am.wav").
                    # NOTE: This section feels a little clunky.
                    name = file.split(".")[0].split("-")
                    if name[-1].lower() == "am" or name[-1].lower() == "pm":
                        name = name[:-1]
                    name = "-".join(name)

                    # Create dictionary of intellectual groups
                    if item["id"] not in items.keys():
                        items[item["id"]] = {}
                    if name not in items[item["id"]].keys():
                        items[item["id"]][name] = []
                    items[item["id"]][name].append(file)

    return items


def check_intellectual_groups(baroqueproject):
    """
    Make sure that intellectual groups are well-formed by:
    (1) Checking that each part number is consecutively numbered.
    (2) Checking that each part only has 6 files.
    (3) Checking that each part has exactly one of the 6 required file formats.
    """
    items = create_intellectual_groups(baroqueproject)

    for item in items.keys():
        # Check that each of the part numbers is consecutive
        part_ids = []
        consecutive = True

        for part_id in items[item].keys():
            part_ids.append(part_id)

        part_ids.sort()

        # NOTE: This logic does not work when files have the following format: 1234-SR-1-1-1.
        for n in range(len(part_ids)):
            if not part_ids[n].endswith(str(n + 1)):
                consecutive = False

        if not consecutive:
            print("ERROR: The parts of item", item, "are not consecutively numbered:", part_ids)

        for part in items[item]:
            # Check that each part only has 6 files.
            if len(items[item][part]) > 6:
                print("ERROR: There are", len(items[item][part]) - 6, "extra files in the", part,
                      "part, which consists of:", items[item][part])

            if len(items[item][part]) < 6:
                print("ERROR: There are", 6 - len(items[item][part]), "missing files in the", part,
                      "part, which consists of:", items[item][part])

            if len(items[item][part]) == 6:
                # Check that the part only has one and only one of each of the required file formats.
                count_formats = {"-am.wav": 0, "-am.wav.md5": 0, "-pm.wav": 0, "-pm.wav.md5": 0, ".mp3": 0,
                                 ".mp3.md5": 0}

                for file in items[item][part]:
                    for format in count_formats.keys():
                        if file.endswith(format):
                            count_formats[format] += 1

                for format in count_formats.keys():
                    if count_formats[format] > 1:
                        print("ERROR: There are", count_formats[format] - 1, "extra", format, "files in the", part,
                              "part:", count_formats)
                    if count_formats[format] < 1:
                        print("ERROR: There is 1 missing", format, "file in the", part, "part:", count_formats)


def validate_file(baroqueproject):
    check_item_files(baroqueproject)
    check_intellectual_groups(baroqueproject)


def validate_structure(baroqueproject, metadata_export):
    """
    If the source directory is a shipment,
    this function runs "validate_directory" on collection/item-level and "validate_file" on item-level.
    If the source directory is a collection, it runs "validate_directory" and "validate_file" on item-level.
    If the source directory is an item, "validate_file" on item-level.
    """
    if baroqueproject.source_type == "shipment":
        levels = ["collections", "items"]
        for level in levels:
            validate_directory(baroqueproject, metadata_export, level)

    elif baroqueproject.source_type == "collection":
        level = "items"
        validate_directory(baroqueproject, metadata_export, level)

    validate_file(baroqueproject)
